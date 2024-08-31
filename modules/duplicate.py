from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from time import sleep

def main():
    # Load Excel workbook for any given "data.xlsx" file
    wb = load_workbook("data.xlsx")
    sheet = wb.active

    # Initialize WebDriver
    service = webdriver.ChromeService(executable_path = "../chromedriver-mac-x64/chromedriver")
    driver = webdriver.Chrome(service = service)
    driver.maximize_window() # For clarity

    # Navigate to pears.io
    driver.get("https://database-example.com")

    # Log in with credentials
    username = "username"
    password = "password"
    driver.find_element(By.ID, "id_email").send_keys(username)
    driver.find_element(By.ID, "id_password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

    # Verify that login is successful
    profile = driver.find_element(By.ID, "account-dropdown")
    assert profile.is_displayed()

    unique = []
    duplicates = []

    # Data entry
    for row in sheet.iter_rows(min_row=2):
        # Navigate to "Partnerships" tab
        driver.get("https://database-example/partnerships/")

        # Retrieve partnership name from data file
        name = row[0].value

        # Search for partnership name
        search = driver.find_element(By.CLASS_NAME, "c-search__input")
        search.clear()
        search.send_keys(name)

        # Check for potential duplicates
        try:
            WebDriverWait(driver, 3).until(EC.presence_of_element_located(
                (By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div[2]/p/p/a")))
            unique.append(name)
            sleep(3)
        except TimeoutException:
            duplicates.append(name)
            sleep(3)
    
    # Print out unique entries for reference
    if unique:
        print("The following unique partnerships were found:")
        for entry in unique:
            print(entry)
    
    print()

    # Print out duplicates for reference
    if duplicates:
        print("The following potential duplicate partnerships were found:")
        for entry in duplicates:
            print(entry)

    # Close WebDriver
    driver.quit()

if __name__ == "__main__":
    main()
