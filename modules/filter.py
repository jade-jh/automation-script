from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import load_workbook

# Initialize WebDriver as global variable
service = webdriver.ChromeService(executable_path = "../chromedriver-mac-x64/chromedriver")
driver = webdriver.Chrome(service = service)

def main():
    # Load Excel workbook for any given "data.xlsx" file
    wb = load_workbook("data/test.xlsx")
    sheet = wb.active

    # Navigate to pears.io
    driver.get("https://database-example.com")

    # Log in with credentials
    username = "username"
    password = "password"
    driver.find_element(By.ID, "id_email").send_keys(username)
    driver.find_element(By.ID, "id_password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

    # Verify that login is successful
    profile = driver.find_element(By.ID, "account-dropdown")
    assert profile.is_displayed()

    # Data entry
    for row in sheet.iter_rows(min_row=2):
        # Navigate to "Partnerships" tab
        driver.get("https://database-example.com/partnerships/")

        # Retrieve partnership name from data file
        name = row[0].value

        # Remove filter for user-created entries
        filter = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
            (By.ID, 'filter-dropdown-Created By')))
        if not "Created By" in filter.text:
            driver.find_element(By.XPATH, "//*[@id='filter-dropdown-Created By']/span").click()

        # Search for partnership name
        search = driver.find_element(By.CLASS_NAME, "c-search__input")
        search.clear()
        search.send_keys(name)

        sleep(3)

    # Close WebDriver
    driver.quit()

if __name__ == "__main__":
    main()
