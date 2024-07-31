from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from time import sleep
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
import threading
import re

# Initialize WebDriver as global variable
driver = webdriver.Chrome(service=webdriver.ChromeService(executable_path = "../chromedriver-mac-x64/chromedriver"))

def scroll(element):
    """
    Scrolls on the current page until the given element is at the top of the viewport.

    Parameters
    ----------
    element : WebElement
        Target element to scroll to.
    
    Returns
    -------
    None
        This function does not return any value.
    """
    driver.execute_script("arguments[0].scrollIntoView({ block: 'start' });", element)
    sleep(1) # Wait for scrolling to finish

def save_and_continue(submit):
    """
    Saves the information entered on the current page and continues on to the next section.

    Parameters
    ----------
    submit : WebElement
        Submit button to be clicked.
    
    Returns
    -------
    None
        This function does not return any value.
    """
    initial_url = driver.current_url
    driver.execute_script("arguments[0].scrollIntoView(true);", submit)
    sleep(1) # Scrolls until the 'Save and Continue' button enters the viewport
    submit.click()
    WebDriverWait(driver, 5).until(EC.url_changes(initial_url)) # Check successful submission

def show_gui(resume_event, network):
    """
    Displays a Tkinter GUI to allow user control over manual input of partnership type/network.

    Parameters
    ----------
    resume_event : Event
        An instance of threading.Event used to signal the completion of any user action.
    network : str
        Network to enter, if applicable; empty otherwise.
    
    Returns
    -------
    None
        This function does not return any value.
    """
    root = tk.Tk()

    # Position GUI window at top right
    width = 200
    height = 100
    x = root.winfo_screenwidth() - width
    y = 0
    root.geometry(f"{width}x{height}+{x}+{y}")

    # Create title, label, and button for GUI window
    if network:
        root.title("Network(s) Reached")
        prompt = tk.Label(root, text="Please select the\n"+network+" network.")
    else:
        root.title("Parternership Type")
        prompt = tk.Label(root, text="Please enter this\npartnership's type.")
    prompt.pack(padx=10, pady=5)
    resume = tk.Button(root, text="I'm done", command=lambda: (resume_event.set(), root.withdraw(), root.destroy()))
    resume.pack(pady=10)

    # Start the event loop
    root.mainloop()

def main():
    """
    Entry point of the script.

    Logs a series of partnerships provided by an Excel file.

    Parameters
    ----------
    None
        This function does not take any values.

    Returns
    -------
    None
        This function does not return any value.
    """
    # Load Excel workbook for a given file
    wb = load_workbook("data/data.xlsx")
    sheet = wb.active

    # Navigate to URL
    driver.get("https://database-example.com")

    # Log in with credentials
    username = "username"
    password = "password"
    driver.find_element(By.ID, "id_email").send_keys(username)
    driver.find_element(By.ID, "id_password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

    # Verify that login is successful
    profile = driver.find_element(By.ID, "account-dropdown")
    assert profile.is_displayed()

    # Duplicate and/or mishandled entries that won't be entered
    duplicates = []
    errors = []

    # Highlights for duplicate and/or mishandled entries on spreadsheet
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Duplicate
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Error: not entered
    orange_fill = PatternFill(start_color="FFA500", end_color = "FFA500", fill_type="solid") # Error: partially completed

    # Data entry
    for row in sheet.iter_rows(min_row=2):
        # Navigate to "Partnerships" tab
        driver.get("https://database-example/partnerships/")

        # Retrieve partnership name from data file
        name = row[0].value

        # Remove filter for user-created entries
        filter = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
            (By.ID, 'filter-dropdown-Created By')))
        if not "Created By" in filter.text:
            driver.find_element(By.XPATH, "//*[@id='filter-dropdown-Created By']/span").click()

        # Search for partnership name
        search = driver.find_element(By.CLASS_NAME, "c-search__input")
        search.clear()
        search.send_keys(name)

        # Check for potential duplicates
        try:
            add_partnership = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
                (By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div[2]/p/p/a")))
            add_partnership.click()
        except TimeoutException:
            duplicates.append(name)
            row[0].fill = yellow_fill # Mark entry as potential duplicate for user review
            continue

        # Begin "General Information"
        try:
            # Enter partnership name
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "id_name"))).send_keys(name)

            # Enter action plan
            action_plan = row[3].value
            ap_prompt = driver.find_element(By.XPATH, "//*[@id='div_id_action_plan']/span")
            scroll(ap_prompt)
            ap_prompt.click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'" + action_plan + "')]"))).click()

            # Enter partnership site by ID
            site = row[5].value
            driver.find_element(By.XPATH, "//*[@id='div_id_site']/span").click()
            driver.find_element(By.XPATH, "/html/body/span/span/span[1]/input").send_keys(site)
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'ID: " + str(site) + "')]"))).click()

            # Enter partnership unit
            unit = row[4].value
            driver.find_element(By.XPATH, "//*[@id='div_id_unit']/span").click()
            driver.find_element(By.XPATH, "/html/body/span/span/span[1]/input").send_keys(unit)
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'" + unit + "')]"))).click()

            # Enter jurisdiction level (CONSTANT)
            driver.find_element(By.XPATH, "//*[@id='div_id_jurisdiction_level']/span").click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'Local')]"))).click()

            # Enter partnership type (with user input)
            resume_event = threading.Event()
            show_gui(resume_event, "")
            sleep(1) # Buffer after button click

            # Enter assistance received
            assistance_received = re.sub(r'\s*\(.*?\)', "", row[12].value).strip().split(",") # Clean string
            assistance_received = [assist.strip() for assist in assistance_received] # Eliminate whitespace
            ar_prompt = driver.find_element(By.XPATH, "//*[@id='div_id_assistance_received']/span")
            scroll(ar_prompt)
            ar_prompt.click()
            enter_ar = driver.find_element(By.XPATH, "//*[@id='div_id_assistance_received']/span/span[1]/span/ul/li/input")
            for assist in assistance_received:
                enter_ar.send_keys(assist)
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                    (By.XPATH, "//li[contains(.,'" + assist + "')]"))).click()

            # Enter assistance provided
            assistance_provided = re.sub(r'\s*\(.*?\)', "", row[13].value).strip().split(",") # Clean string
            assistance_provided = [assist.strip() for assist in assistance_provided] # Eliminate whitespace
            driver.find_element(By.XPATH, "//*[@id='div_id_assistance_provided']/span").click()
            enter_ap = driver.find_element(By.XPATH, "//*[@id='div_id_assistance_provided']/span/span[1]/span/ul/li/input")
            for assist in assistance_provided:
                enter_ap.send_keys(assist)
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                    (By.XPATH, "//li[contains(.,'" + assist + "')]"))).click()
            
            # Enter whether the partner received direct funding
            funding = row[14].value
            funding_prompt = driver.find_element(By.XPATH, "//*[@id='div_id_is_snaped_funded']/span")
            scroll(funding_prompt)
            funding_prompt.click()
            enter_funding = driver.find_element(By.XPATH, "/html/body/span/span/span[1]/input")
            enter_funding.send_keys(funding)
            enter_funding.send_keys(Keys.ENTER) # For Element Click Intercepted Exception

            # Enter relevant intervention types
            direct_ed = row[15].value # Consider direct education alone
            if direct_ed == 1:
                driver.find_element(By.ID, "id_intervention_types_0").click()
            
            # Enter program activity comments
            comments = row[23].value
            if comments:
                comments_prompt = driver.find_element(By.CSS_SELECTOR, "div.fr-element[contenteditable='true']")
                driver.execute_script("arguments[0].innerHTML = arguments[1];", comments_prompt, comments)

            # Submit "General Information"
            submit_gen_info = driver.find_element(By.XPATH, "//*[@id='form_id']/button[1]")
            save_and_continue(submit_gen_info)
        
        except Exception as e:
            errors.append(name)
            row[0].fill = red_fill # Mark entry as unentered for user review
            print(f"An error occurred while entering {name}: {e}")
            continue

        try:
            # Begin "Collaborators"
            if row[16].value:
                collaborators = [collab.strip() for collab in row[16].value.split(",")]
                for collab in collaborators:
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[@id='app']/div/div[2]/div[1]/div/div[1]/button"))).click()
                    enter_user = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.CSS_SELECTOR, "div.vs__selected-options input.vs__search")))
                    enter_user.send_keys(collab)
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.XPATH, "//span[contains(.,'" + collab + "')]"))).click()
                    driver.find_element(By.XPATH, "//*[@id='id_Contributor']").click()
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.CSS_SELECTOR, "#id_Access"))).click()
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.XPATH, "//li[contains(.,'View & Edit')]"))).click() # CONSTANT
                    driver.find_element(By.XPATH, "//*[@id='app']/div/div[2]/div[1]/div[2]/div/div/div/div[3]/button[1]").click()
                sleep(1) # Buffer to allow page to load

            # Remove user as collaborator
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//*[@id='app']/div/div[2]/div[1]/div/div[2]/table/tbody/tr/td[5]/div/a[2]"))).click()
            sleep(1) # Buffer to allow page to load
        
            # Submit "Collaborators"
            submit_collab = driver.find_element(By.XPATH, "//*[@id='main-content']/form/button[1]")
            save_and_continue(submit_collab)

            # Begin "Custom Data"
            # Enter grant goals
            if not row[17].value: # Check if cell is empty
                errors.append(name)
                row[0].fill = orange_fill # Mark entry as incomplete for user review
                continue
            goals = row[17].value.split(",")
            enter_goals = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                ((By.XPATH, "//*[@id='vs1__combobox']/div[1]/input"))))
            for goal in goals:
                enter_goals.send_keys(goal)
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                    (By.XPATH, "//li[contains(.,'" + goal + "')]"))).click()

            # Enter network reached
            network = row[8].value
            resume_event = threading.Event() # Create event object for synchronization
            show_gui(resume_event, network)
            sleep(1) # Buffer after button click

            # Enter special projects (CONSTANT)
            projects = driver.find_element(By.XPATH, "//*[@id='vs3__combobox']/div[1]/input")
            driver.execute_script("arguments[0].scrollIntoView(true);", projects)
            sleep(1)
            projects.click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'None')]"))).click()

            # Submit "Custom Data"
            submit_custom_data = driver.find_element(By.XPATH, "//*[@id='app']/div/div[2]/div[1]/div[2]/button[1]")
            save_and_continue(submit_custom_data)

            # Begin "Evaluation"
            # Enter relationship depth
            relationship = row[18].value
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//*[@id='div_id_relationship_depth']/span"))).click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'" + relationship + "')]"))).click()

            # Enter assessment tool (CONSTANT - Expecting "None")
            tool = row[19].value
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//*[@id='div_id_assessment_tool']/span"))).click()
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//li[contains(.,'" + tool + "')]"))).click()
            
            # Enter partnership accomplishments
            accomplishment = row[20].value
            enter_accom = driver.find_element(By.ID, "id_accomplishments")
            scroll(enter_accom)
            enter_accom.send_keys(accomplishment)

            # Enter lessons learned
            lessons = row[21].value
            enter_lessons = driver.find_element(By.ID, "id_lessons_learned")
            enter_lessons.send_keys(lessons)

            # Submit "Evaluation"
            submit_eval = driver.find_element(By.XPATH, "//*[@id='main-content']/div[2]/div[1]/form/button[1]")
            save_and_continue(submit_eval)

            # Begin "Meetings & Events"
            # Enter meetings/events (CONSTANT)
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//*[@id='div_id_has_events']/span"))).click()
            enter_me = driver.find_element(By.XPATH, "/html/body/span/span/span[1]/input")
            enter_me.send_keys("No")
            enter_me.send_keys(Keys.ENTER)
        
            # Submit "Meetings & Events"
            submit_meetings_events = driver.find_element(By.XPATH, "//*[@id='main-content']/div[2]/div[1]/form/button[1]")
            save_and_continue(submit_meetings_events)
        
        except Exception as e:
            errors.append(name)
            row[0].fill = orange_fill # Mark entry as incomplete for user review
            print(f"An error occurred while entering {name}: {e}")
            continue

    # Print out duplicates for reference
    if duplicates:
        print("The following potential duplicate partnerships were not entered:")
        for duplicate in duplicates:
            print(duplicate)
    
    if duplicates and errors:
        print() # Line break

    # Print out errors for reference
    if errors:
        print("One or more errors occurred while attempting to enter the following partnerships:")
        for error in errors:
            print(error)

    # Save workbook (with highlights for duplicates/errors) to new file
    wb.save("post_entry_data.xlsx")
    print("\nSaved overview as 'post_entry_data.xlsx'")

    # Close WebDriver
    driver.quit()

if __name__ == "__main__":
    main()